# <editor-fold desc="Imports"
import unittest
import openpyxl as excelReader
from selenium import webdriver
from pyunitreport import HTMLTestRunner
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
# </editor-fold>

# <editor-fold desc="So Variables"
excelFilePath = (r"C:\Users\Amr\Desktop\testData.xlsx")
current_URL = "http://stageavl.afaqy.sa/"
chromePath = (r"Z:\AfaqyAutomationPython\chromedriver.exe")
driver = webdriver.Chrome
# </editor-fold>

class TestAddingUnit(unittest.TestCase):
    # <editor-fold desc="Step 0 - Prepare test data"
    def test_start(self):
        excelTestData = excelReader.load_workbook(excelFilePath)
        activeWorkBook = excelTestData.active
        rowsLen = activeWorkBook.max_row
        print("Rows Found: " + str(rowsLen - 1))
        #input("Press Enter to start...")

        self.startDriver()

        self.loginToAccount()

        self.openUnitsModule()

        for i in range(2, rowsLen + 1, 1):
            testDataDictionary = self.readFullRow(i, activeWorkBook)

            addUnitModalDialogElement = self.openAddUnitModal()

            self.mainTabProcess(addUnitModalDialogElement, testDataDictionary)

            self.profileTabProcess(addUnitModalDialogElement, testDataDictionary)

            self.groupTabProcess(addUnitModalDialogElement, testDataDictionary)

            self.fuelTabProcess(addUnitModalDialogElement)

            self.sensorTabProcess(addUnitModalDialogElement, testDataDictionary)

            self.commandsTabProcess(addUnitModalDialogElement, testDataDictionary)

            self.submitAddingUnit(addUnitModalDialogElement)

        print("Test Completed")

    def readFullRow(self, targetRowIndex, activeWorkBook):
        columnLen = activeWorkBook.max_column
        testDataDictionary = {}
        for i in range(1, columnLen + 1):
            cellHeader = activeWorkBook.cell(row=1, column=i)
            cellValue = activeWorkBook.cell(row=targetRowIndex, column=i)
            if cellValue.value != None:
                testDataDictionary.update({cellHeader.value: cellValue.value})
        return testDataDictionary
    # </editor-fold>

    # - Step 1 - Navigate to web page
    def startDriver(self):
        global driver
        driver = webdriver.Chrome(executable_path=chromePath)
        driver.get(current_URL)
        loginForm = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, "form-validation")))
        self.assertIsNotNone(loginForm)

    # - Step 2 - Login
    def loginToAccount(self):
        mailTextBox = driver.find_element(By.CSS_SELECTOR, "input[id='email']")
        mailTextBox.send_keys("Rawahel")

        passwordTextBox = driver.find_element(By.CSS_SELECTOR, "input[id='password']")
        passwordTextBox.send_keys("Asd@123")

        submitButton = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
        submitButton.click()

        headerNotchElement = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "headerNotch")))
        self.assertIsNotNone(headerNotchElement)
        #--- Assert here by checking the login form visibility

    # - Step 3 - Units
    def openUnitsModule(self):
        sideMenu = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "div[id='headerNotch']" + '>' + "button[type='button']")))
        sideMenu.click()

        sideMenuScope = driver.find_element(By.TAG_NAME, "afaqy-menu")
        unitsPage = WebDriverWait(sideMenuScope, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href^='/units']")))
        unitsPage.click()
        self.assertIn( "- Units", driver.title)
        # --- Assert here by checking the web page title

    # - Step 4 - Click on Add
    def openAddUnitModal(self):
        pageHeader = driver.find_element(By.CSS_SELECTOR, "div[id='page-headers']")
        addUnitButton = pageHeader.find_element(By.CSS_SELECTOR, "a[href='/units/add']")
        addUnitButton.click()

        addUnitDialog = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.TAG_NAME, "afaqy-modal")))
        dialogTitle = addUnitDialog.find_element(By.TAG_NAME, "h5").get_attribute("innerText")

        self.assertEqual(dialogTitle, "Add")
        return addUnitDialog
        # --- Assert here by checking the dialog title = (Add)

    # - Step 5 - Main Tab
    def mainTabProcess(self, addUnitDialogElement, testDataDictionary):
        unitNameTextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='name']")
        unitNameTextBox.send_keys(testDataDictionary["Unit Name"])

        deviceIMEITextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='imei']")
        deviceIMEITextBox.send_keys(testDataDictionary["IMEI"])

        deviceSerialTextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='device_serial']")
        deviceSerialTextBox.send_keys(testDataDictionary["Device Serial"])

        deviceSIMTextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='sim_number']")
        deviceSIMTextBox.send_keys(testDataDictionary["Sim Number"])

        simSerialTextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='sim_serial']")
        simSerialTextBox.send_keys(testDataDictionary["Sim Serial"])

        untiDeviceDropBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "select[testid='device']")
        untiDeviceDropBox.click()
        deviceValue = untiDeviceDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Device Protocol'].lower()}']")
        deviceValue.click()
        # --- No Assert here

    # - Step 6 - Profile Tab : Tab index = 2
    def profileTabProcess(self, addUnitDialogElement, testDataDictionary):
        self.switchTab(addUnitDialogElement, 2)
        # ======================
        unitPlateTextBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "input[testid='plate_number']")
        unitPlateTextBox.send_keys(testDataDictionary["Plate Number"])

        untiFuelDropBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "select[testid='fuel_type']")
        untiFuelDropBox.click()
        fuelValue = untiFuelDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Fuel Type']}']")
        fuelValue.click()

        untiTrailerDropBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "afaqy-custom-select[testid='tailer_id']")
        dropBoxInputBox = untiTrailerDropBox.find_element(By.CSS_SELECTOR, "input[type='text']")
        dropBoxInputBox.send_keys(testDataDictionary["Trailer ID"])

        dropBoxItemsPanel = untiTrailerDropBox.find_element(By.TAG_NAME, "ng-dropdown-panel")
        dropBoxDriversList = dropBoxItemsPanel.find_elements(By.CSS_SELECTOR, "div[role='option']")
        while len(dropBoxDriversList) > 1:
            for x, driverName in enumerate(dropBoxDriversList[1:]):
                if driverName.get_attribute("innerText") == testDataDictionary["Trailer ID"]:
                    driverName.click()
                    break
            break

        unitDriverDropBox = addUnitDialogElement.find_element(By.CSS_SELECTOR, "afaqy-custom-select[testid='driver_id']")
        dropBoxInputBox = unitDriverDropBox.find_element(By.CSS_SELECTOR, "input[type='text']")
        dropBoxInputBox.send_keys(testDataDictionary["Driver Name"])

        dropBoxItemsPanel = unitDriverDropBox.find_element(By.TAG_NAME, "ng-dropdown-panel")
        dropBoxDriversList = dropBoxItemsPanel.find_elements(By.CSS_SELECTOR, "div[role='option']")
        while len(dropBoxDriversList) > 1:
            for x, driverName in enumerate(dropBoxDriversList[1:]):
                if driverName.get_attribute("innerText") == testDataDictionary["Driver Name"]:
                    driverName.click()
                    break
            break
        # --- No Assert here

    # - Step 7 - Groups Tab : Tab index = 3
    def groupTabProcess(self, addUnitDialogElement, testDataDictionary):
        self.switchTab(addUnitDialogElement, 3)
        # ======================
        activeTabArea = addUnitDialogElement.find_element(By.CSS_SELECTOR, "tab[class='tab-pane active']")
        unitGroupSearchBox = activeTabArea.find_element(By.CSS_SELECTOR, "input[type='text']")
        unitGroupSearchBox.send_keys(testDataDictionary["Group Name"])

        groupsGridList = addUnitDialogElement.find_element(By.CSS_SELECTOR, "div[class='flexgridSelect']")
        groupsGridRowsSpace = groupsGridList.find_element(By.CSS_SELECTOR, "div[wj-part='cells']")
        groupsRowsList = groupsGridRowsSpace.find_elements(By.CSS_SELECTOR, "div[class='wj-row']")
        while len(groupsRowsList) > 1:
            for x, groupName in enumerate(groupsRowsList[1:]):
                if groupName.get_attribute("innerText") == testDataDictionary["Group Name"]:
                    groupName.find_element(By.CSS_SELECTOR, "div[role='gridcell']").click()
                    break
            break
        buttonsGroupDiv = addUnitDialogElement.find_element(By.CSS_SELECTOR, "div[class*='assign-buttons']")
        addSingleGroupButton = buttonsGroupDiv.find_element(By.CSS_SELECTOR, "button:nth-of-type(2)")
        addSingleGroupButton.click()
        # --- No Assert here

    # - Step 8 - Fuel Consumption : Tab index = 3
    def fuelTabProcess(self, addUnitDialogElement):
        self.switchTab(addUnitDialogElement, 4)
        # ======================
        addSingleGroupButton = addUnitDialogElement.find_element(By.CSS_SELECTOR, "label[for='unit-fc-math-enable']")
        addSingleGroupButton.click()
        # --- No Assert here

    # - Step 9 - Sensor Tab : Tab index = 7
    def sensorTabProcess(self, addUnitDialogElement, testDataDictionary):
        self.switchTab(addUnitDialogElement, 7)
        # ======================
        activeTabArea = addUnitDialogElement.find_element(By.CSS_SELECTOR, "tab[class='tab-pane active']")
        addSensorButton = activeTabArea.find_element(By.TAG_NAME, "button")
        addSensorButton.click()

        addSensorDialog = addUnitDialogElement.find_element(By.TAG_NAME, "unit-sensor-form")

        sensorNameTextBox = addSensorDialog.find_element(By.CSS_SELECTOR, "input[formcontrolname='name']")
        sensorNameTextBox.send_keys(testDataDictionary["Sensor Name"])

        sensorTypeDropBox = addSensorDialog.find_element(By.CSS_SELECTOR, "select[formcontrolname='type']")
        sensorTypeDropBox.click()
        typeValue = sensorTypeDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Type'].lower()}']")
        typeValue.click()

        sensorParamDropBox = addSensorDialog.find_element(By.CSS_SELECTOR, "select[formcontrolname='param']")
        sensorParamDropBox.click()
        paramValue = sensorParamDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Param'].lower()}']")
        paramValue.click()

        sensorShowToolTip = addSensorDialog.find_element(By.CSS_SELECTOR, "label[for='tooltip_show']")
        sensorShowToolTip.click()

        sensorMeasurmentDropBox = addSensorDialog.find_element(By.CSS_SELECTOR, "select[formcontrolname='result_type']")
        sensorMeasurmentDropBox.click()
        measurmentValue = sensorMeasurmentDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Measure Type'].lower()}']")
        measurmentValue.click()

        smallModelFooter = addSensorDialog.find_element(By.TAG_NAME, "form-actions")
        saveButton = smallModelFooter.find_element(By.CSS_SELECTOR, "button:nth-of-type(1)")
        saveButton.click()
        # --- No Assert here

    # - Step 9 - Commands Tab : Tab index = 8
    def commandsTabProcess(self, addUnitDialogElement, testDataDictionary):
        self.switchTab(addUnitDialogElement, 8)
        # ======================
        activeTabArea = addUnitDialogElement.find_element(By.CSS_SELECTOR, "tab[class='tab-pane active']")
        addCommandsButton = activeTabArea.find_element(By.TAG_NAME, "button")
        addCommandsButton.click()

        addCommandsDialog = addUnitDialogElement.find_element(By.TAG_NAME, "unit-commands-form")

        commandNameTextBox = addCommandsDialog.find_element(By.CSS_SELECTOR, "input[formcontrolname='name']")
        commandNameTextBox.send_keys(testDataDictionary["Command Name"])

        commandTypeDropBox = addCommandsDialog.find_element(By.CSS_SELECTOR,
                                                            "select[formcontrolname='protocol_command_id']")
        commandTypeDropBox.click()
        typeValue = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, f"option[value='{testDataDictionary['Type'].lower()}']")))
        typeValue.click()

        commandChannelDropBox = addCommandsDialog.find_element(By.CSS_SELECTOR, "select[formcontrolname='channel']")
        commandChannelDropBox.click()
        typeValue = commandChannelDropBox.find_element(By.CSS_SELECTOR, f"option[value='{testDataDictionary['Channel'].lower()}']")
        typeValue.click()

        commandNameTextBox = addCommandsDialog.find_element(By.CSS_SELECTOR, "input[formcontrolname='param']")
        commandNameTextBox.send_keys(testDataDictionary["Message"])

        smallModelFooter = addCommandsDialog.find_element(By.TAG_NAME, "form-actions")
        saveButton = smallModelFooter.find_element(By.CSS_SELECTOR, "button:nth-of-type(1)")
        saveButton.click()

    # - Step 10 - Submit Adding Unit
    def submitAddingUnit(self, addUnitDialogElement):
        bigModelFooter = addUnitDialogElement.find_element(By.TAG_NAME, "form-actions")
        saveButton = bigModelFooter.find_element(By.CSS_SELECTOR, "button:nth-of-type(1)")
        saveButton.click()

        notificationDivision = driver.find_element(By.TAG_NAME, "simple-notifications")
        sucessNotification = WebDriverWait(notificationDivision, 60).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[class*='success']")))

        self.assertIsNotNone(sucessNotification)

    # - External Step - Switch Tab Function
    def switchTab(self, addUnitDialogElement, tabIndex):
        tabLists = addUnitDialogElement.find_element(By.CSS_SELECTOR, "ul[class='nav nav-tabs']")
        targetTap = tabLists.find_element(By.CSS_SELECTOR,
                                          f"li:nth-of-type({tabIndex})")  # ---- needed tab index (Needed tab)
        targetTap.click()


if __name__ == '__main__':
   unittest.main(testRunner=HTMLTestRunner(output="Reports"))
